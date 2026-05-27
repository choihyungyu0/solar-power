from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


SCRIPT_PATH = Path(__file__).resolve()
BACKEND_ROOT = SCRIPT_PATH.parents[1]
REPO_ROOT = SCRIPT_PATH.parents[3]
EXCEL_PATH = REPO_ROOT / "data" / "policy" / "태양광_지원사업_정리.xlsx"

sys.path.insert(0, str(BACKEND_ROOT))

from app.subsidy_rag import (  # noqa: E402
    build_subsidy_chunk_text,
    generate_embedding,
    insert_subsidy_chunk,
    upsert_subsidy_document,
)
from app.supabase_client import deactivate_subsidy_rag_source  # noqa: E402


SOURCE_TITLE = "태양광 지원사업 정리"
MAIN_PROGRAM_NAME = "경기 주택태양광 지원사업"
DEFAULT_SOURCE_YEAR = 2026
CITY_SUBSIDY_SHEET = "02_경기주택태양광_시군별"
SKIPPED_SHEET_PREFIXES = ("00_",)


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        normalized = (
            value.replace(",", "")
            .replace("만원", "0000")
            .replace("원", "")
            .replace("%", "")
            .strip()
        )

        if not normalized:
            return None

        try:
            return float(normalized)
        except ValueError:
            return None

    return None


def _int(value: Any) -> int | None:
    number = _number(value)

    return round(number) if number is not None else None


def _get(row: dict[str, Any], *headers: str) -> Any:
    wanted = {_normalize_key(header) for header in headers}

    for key, value in row.items():
        if key.startswith("__"):
            continue

        if _normalize_key(key) in wanted:
            return value

    return None


def _pick(row: dict[str, Any], *keywords: str) -> Any:
    for key, value in row.items():
        if key.startswith("__"):
            continue

        normalized_key = key.replace(" ", "").lower()

        if any(keyword.replace(" ", "").lower() in normalized_key for keyword in keywords):
            return value

    return None


def _first_year(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text)

    if not match:
        return None

    year = int(match.group(1))
    return year if 2000 <= year <= 2100 else None


def _extract_year(row: dict[str, Any]) -> int:
    explicit = _int(_pick(row, "연도", "년도", "year"))

    if explicit and 2000 <= explicit <= 2100:
        return explicit

    year_from_text = _first_year(" ".join(_text(value) for value in row.values()))

    if year_from_text:
        return year_from_text

    return DEFAULT_SOURCE_YEAR or datetime.now().year


def _normalize_rate(value: float | None) -> float | None:
    if value is None:
        return None

    return round(value * 100, 2) if 0 < value <= 1 else round(value, 2)


def _first_percent(text: str) -> float | None:
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)

    if match:
        return _normalize_rate(float(match.group(1)))

    number = _number(text)
    return _normalize_rate(number)


def _krw_values_from_text(text: str) -> list[int]:
    values: list[int] = []

    for match in re.finditer(r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*천원", text):
        values.append(round(float(match.group(1).replace(",", "")) * 1000))

    for match in re.finditer(r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*만원", text):
        values.append(round(float(match.group(1).replace(",", "")) * 10000))

    for match in re.finditer(r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*원", text):
        before = text[max(0, match.start() - 1) : match.start()]
        if before in {"천", "만"}:
            continue

        values.append(round(float(match.group(1).replace(",", ""))))

    return values


def _krw_after_keyword(text: str, keywords: tuple[str, ...]) -> int | None:
    for keyword in keywords:
        index = text.find(keyword)

        if index < 0:
            continue

        values = _krw_values_from_text(text[index:])

        if values:
            return values[0]

    values = _krw_values_from_text(text)
    return values[0] if values else None


def _thousand_krw(value: Any) -> int | None:
    number = _number(value)
    return round(number * 1000) if number is not None else None


def _normalize_program_name(row: dict[str, Any]) -> str:
    program_name = _text(_get(row, "지원사업명", "사업명", "program_name", "program"))

    if "경기" in program_name and "태양광" in program_name:
        return MAIN_PROGRAM_NAME

    return program_name or MAIN_PROGRAM_NAME


def _normalize_region_sido(row: dict[str, Any]) -> str:
    value = _text(_get(row, "시도", "광역", "region_sido", "sido"))

    if value:
        return "경기도" if value in {"경기", "경기도"} else value

    primary_text = " ".join(
        _text(value)
        for key, value in row.items()
        if key in {"지원사업명", "운영주체", "지원대상", "__sheet_heading"}
    )

    if "경기" in primary_text or "화성" in primary_text:
        return "경기도"

    if "한국에너지공단" in primary_text or "전국" in primary_text or "국비" in primary_text:
        return "전국"

    return ""


def _normalize_region_sigungu(row: dict[str, Any]) -> str | None:
    value = _text(_get(row, "시군구", "시군", "시군명", "지역", "지자체", "sigungu"))

    if value and value not in {"경기도", "경기"}:
        return value

    for key in ("지원사업명", "운영주체", "지원대상", "비고(출처·중복)"):
        cell = row.get(key)
        text = _text(cell)

        if "화성시" in text:
            return "화성시"

        if text.endswith("시") or text.endswith("군") or text.endswith("구"):
            return text

    return None


def _is_stacking_allowed(row: dict[str, Any], program_name: str) -> bool | None:
    stacking_text = " ".join(
        _text(value)
        for key, value in row.items()
        if key in {"비고(출처·중복)", "지원자격(조건)", "지원내용", "신청방법·시기", "eligibility_note"}
    )

    blocked_words = ("중복불가", "중복 불가", "택일", "제외", "미지원", "안됨", "금지")

    if any(word in stacking_text for word in blocked_words):
        return False

    if program_name == MAIN_PROGRAM_NAME:
        return False

    return None


def _row_to_city_subsidy_chunk(row: dict[str, Any], chunk_index: int) -> dict[str, Any]:
    region_sigungu = _text(_get(row, "시군명"))
    province_subsidy = _thousand_krw(_get(row, "도 지원금(천원)")) or 0
    city_subsidy = _thousand_krw(_get(row, "시군 보조금(천원)")) or 0
    self_payment = _thousand_krw(_get(row, "자부담금(천원)"))
    total_subsidy = province_subsidy + city_subsidy if province_subsidy or city_subsidy else None
    support_households = _int(_get(row, "지원 가구수"))
    total_rate = _normalize_rate(_number(_get(row, "총 지원율(도+시군)")))
    city_rate = _normalize_rate(_number(_get(row, "시군 보조금 비율")))

    note_parts = [
        f"도 지원금 {province_subsidy:,}원" if province_subsidy else "",
        f"시군 보조금 {city_subsidy:,}원" if city_subsidy else "",
        f"자부담 {self_payment:,}원" if self_payment is not None else "",
        f"지원 가구수 {support_households:,}가구" if support_households is not None else "",
        f"시군 보조율 {city_rate:g}%" if city_rate is not None else "",
        f"총 지원율 {total_rate:g}%" if total_rate is not None else "",
        "실제 접수 가능 여부와 예산 잔여액은 해당 연도 공고 확인 필요",
    ]

    chunk = {
        "chunk_index": chunk_index,
        "chunk_type": "xlsx-city-subsidy-row",
        "region_sido": "경기도",
        "region_sigungu": region_sigungu or None,
        "program_name": MAIN_PROGRAM_NAME,
        "target_building_type": "경기도 소재 주택태양광 3kW 설치 검토 대상",
        "subsidy_amount_krw": total_subsidy,
        "subsidy_rate": total_rate,
        "max_subsidy_krw": total_subsidy,
        "self_payment_krw": self_payment,
        "stacking_allowed": False,
        "eligibility_note": " · ".join(part for part in note_parts if part),
        "source_title": f"{SOURCE_TITLE} - {CITY_SUBSIDY_SHEET}",
        "source_url": None,
        "source_year": _extract_year(row),
        "raw_payload": {
            key: _text(value)
            for key, value in row.items()
            if not key.startswith("__") and _text(value)
        },
        "is_active": True,
        "is_test": False,
    }
    chunk["chunk_text"] = build_subsidy_chunk_text(chunk)

    return chunk


def _row_to_chunk(sheet_name: str, row: dict[str, Any], chunk_index: int) -> dict[str, Any]:
    if sheet_name == CITY_SUBSIDY_SHEET:
        return _row_to_city_subsidy_chunk(row, chunk_index)

    program_name = _normalize_program_name(row)
    region_sido = _normalize_region_sido(row) or "경기도"
    region_sigungu = _normalize_region_sigungu(row)
    target_building_type = _text(_get(row, "지원대상", "target_building_type", "target")) or "주택/공동주택 검토"
    support_limit = _text(_get(row, "지원한도\n(설치비 대비)", "지원한도(설치비 대비)", "지원한도"))
    support_content = _text(_get(row, "지원내용"))
    note = _text(_get(row, "비고(출처·중복)", "비고"))
    joined_policy_text = " ".join(part for part in (support_limit, support_content, note) if part)
    subsidy_amount = _krw_after_keyword(joined_policy_text, ("지원금", "보조금", "보조", "시비", "도비", "국비"))
    max_subsidy = _krw_after_keyword(support_limit, ("최대", "한도", "상한"))
    self_payment = _krw_after_keyword(note, ("자부담", "본인부담"))
    subsidy_rate = _first_percent(support_limit)
    stacking_allowed = _is_stacking_allowed(row, program_name)
    eligibility_parts = [
        _text(_get(row, "지원자격(조건)", "지원자격", "조건")),
        _text(_get(row, "신청방법·시기", "신청방법", "시기")),
        note,
    ]

    source_year = _extract_year(row)
    chunk = {
        "chunk_index": chunk_index,
        "chunk_type": "xlsx-row",
        "region_sido": region_sido,
        "region_sigungu": region_sigungu,
        "program_name": program_name,
        "target_building_type": target_building_type,
        "subsidy_amount_krw": subsidy_amount,
        "subsidy_rate": subsidy_rate,
        "max_subsidy_krw": max_subsidy,
        "self_payment_krw": self_payment,
        "stacking_allowed": stacking_allowed,
        "eligibility_note": " · ".join(part for part in eligibility_parts if part),
        "source_title": f"{SOURCE_TITLE} - {sheet_name}",
        "source_url": _text(_get(row, "확인 링크", "source_url")) or None,
        "source_year": source_year,
        "raw_payload": {
            key: _text(value)
            for key, value in row.items()
            if not key.startswith("__") and _text(value)
        },
        "is_active": True,
        "is_test": False,
    }
    chunk["chunk_text"] = build_subsidy_chunk_text(chunk)

    return chunk


def _find_header_index(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        normalized_cells = {_normalize_key(_text(cell)) for cell in row if _text(cell)}

        if "no" in normalized_cells and (
            "지원사업명" in normalized_cells or "시군명" in normalized_cells
        ):
            return index

    return None


def _load_rows() -> list[tuple[str, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is required. Run pip install -r requirements.txt in services/climate_backend.")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print("Place 태양광_지원사업_정리.xlsx at data/policy/")
        sys.exit(1)

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rows: list[tuple[str, dict[str, Any]]] = []

    for worksheet in workbook.worksheets:
        if worksheet.title.startswith(SKIPPED_SHEET_PREFIXES):
            continue

        raw_rows = list(worksheet.iter_rows(values_only=True))
        header_index = _find_header_index(raw_rows)

        if header_index is None:
            continue

        headers = raw_rows[header_index]

        if not headers:
            continue

        normalized_headers = [
            _text(header) or f"column_{index + 1}"
            for index, header in enumerate(headers)
        ]
        heading_text = " / ".join(
            " ".join(_text(cell) for cell in raw_row if _text(cell))
            for raw_row in raw_rows[:header_index]
            if any(_text(cell) for cell in raw_row)
        )

        for row_number, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
            row = {
                normalized_headers[index]: raw_row[index] if index < len(raw_row) else None
                for index in range(len(normalized_headers))
            }

            if not any(_text(value) for value in row.values()):
                continue

            if _int(_get(row, "No")) is None:
                continue

            row["__sheet_name"] = worksheet.title
            row["__sheet_heading"] = heading_text
            row["__excel_row_number"] = row_number
            rows.append((worksheet.title, row))

    return rows


def main() -> int:
    if not os.getenv("OPENAI_API_KEY", "").strip():
        print("OPENAI_API_KEY is required in the backend environment. The key value will not be printed.")
        return 1

    if os.getenv("ENABLE_SUBSIDY_RAG", "").strip().lower() != "true":
        print("ENABLE_SUBSIDY_RAG=true is required to seed embeddings.")
        return 1

    rows = _load_rows()
    documents_by_sheet: dict[str, str] = {}
    documents_inserted = 0
    chunks_inserted = 0
    failed_rows = 0
    deactivate_result = deactivate_subsidy_rag_source(f"{SOURCE_TITLE} -")

    if deactivate_result.get("ok") is not True:
        print(
            {
                "documentsInserted": 0,
                "chunksInserted": 0,
                "failedRows": len(rows),
                "error": "Failed to deactivate existing subsidy RAG rows.",
                "errorType": deactivate_result.get("errorType"),
            }
        )
        return 1

    for index, (sheet_name, row) in enumerate(rows):
        try:
            if sheet_name not in documents_by_sheet:
                document_result = upsert_subsidy_document(
                    {
                        "source_type": "xlsx",
                        "source_title": f"{SOURCE_TITLE} - {sheet_name}",
                        "source_url": None,
                        "source_year": _extract_year(row),
                        "region_sido": _normalize_region_sido(row) or "경기도",
                        "region_sigungu": _normalize_region_sigungu(row),
                        "program_name": _normalize_program_name(row),
                        "document_version": datetime.now().strftime("%Y%m%d"),
                        "raw_metadata": {
                            "sheetName": sheet_name,
                            "sourcePath": "data/policy/태양광_지원사업_정리.xlsx",
                        },
                        "is_active": True,
                        "is_test": False,
                    }
                )

                if document_result.get("ok") is not True or not isinstance(document_result.get("id"), str):
                    failed_rows += 1
                    continue

                documents_by_sheet[sheet_name] = document_result["id"]
                documents_inserted += 1

            chunk = _row_to_chunk(sheet_name, row, index)
            embedding_result = generate_embedding(chunk["chunk_text"])

            if embedding_result.get("ok") is not True:
                failed_rows += 1
                continue

            chunk["embedding"] = embedding_result["embedding"]
            chunk_result = insert_subsidy_chunk(documents_by_sheet[sheet_name], chunk)

            if chunk_result.get("ok") is True:
                chunks_inserted += 1
            else:
                failed_rows += 1
        except Exception:
            failed_rows += 1

    print(
        {
            "documentsDeactivated": deactivate_result.get("documentsDeactivated", 0),
            "chunksDeactivated": deactivate_result.get("chunksDeactivated", 0),
            "documentsInserted": documents_inserted,
            "chunksInserted": chunks_inserted,
            "failedRows": failed_rows,
        }
    )

    return 0 if chunks_inserted > 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
